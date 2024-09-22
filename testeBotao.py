import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Carregar a tabela de referência
referencias_df = pd.read_excel('TabelaRelacionada.xlsx', sheet_name='Calibração')

# Criar um DataFrame vazio para armazenar os resultados
resultados = pd.DataFrame()

# Percorrer todos os arquivos .xlsx na pasta
for arquivo in os.listdir('.'):
    if arquivo.endswith('.xlsx') and arquivo != 'TabelaRelacionada.xlsx':
        # Carregar a planilha 'Calibração'
        try:
            itens_df = pd.read_excel(arquivo, sheet_name='Calibração')

            # Fazer o merge das DataFrames com base na coluna 'Item'
            merged_df = pd.merge(itens_df, referencias_df[['Item', 'Validade']], on='Item', how='left')

            # Adicionar uma coluna 'Arquivo' com o nome do arquivo
            merged_df['Arquivo'] = arquivo

            # Adicionar uma coluna 'Status' para verificar se a validade do item está abaixo da validade limite
            merged_df['Status'] = merged_df.apply(
                lambda row: 'Abaixo da Validade' if row['Validade_x'] < row['Validade_y'] else 'Dentro da Validade',
                axis=1
            )

            # Filtrar apenas os itens abaixo da validade
            itens_above_limit = merged_df[merged_df['Status'] == 'Abaixo da Validade']

            # Selecionar apenas as colunas necessárias
            resultados = pd.concat([resultados, itens_above_limit[['Item', 'Validade_x', 'Validade_y', 'Arquivo', 'Status']]])

        except Exception as e:
            print(f"Erro ao processar {arquivo}: {e}")

# Salvar o resultado em um novo arquivo Excel
resultado_path = 'resultado_comparacao.xlsx'
resultados.to_excel(resultado_path, index=False)

# Ajustar a largura das colunas usando openpyxl
wb = load_workbook(resultado_path)
ws = wb.active

# Adicionar hiperlinks na coluna F para cada item fora da validade
for row in range(2, len(resultados) + 2):  # +2 para considerar cabeçalho e índice 1
    arquivo_nome = resultados.iloc[row - 2]['Arquivo']
    item_nome = resultados.iloc[row - 2]['Item']
    
    # Criar o hiperlink para a planilha de origem
    ws[f'F{row}'] = f'Acessar {item_nome}'
    ws[f'F{row}'].hyperlink = f'{arquivo_nome}#Calibração'

# Ajustar a largura das colunas
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

wb.save(resultado_path)
wb.close()

print("Comparação completa. O resultado foi salvo em 'resultado_comparacao.xlsx'.")
